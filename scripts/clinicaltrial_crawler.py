"""
Author: Cs.Zhang, @2018.02.06
Usage: python script.py data_table
Describe: A Python3 script for 'clinicaltrials.org' search result downloading and parsing.
Input: table of cancer + gene
Output: table with rownames ([1]Conditions [2]Gene [3]Interventions [4]Title [5]Phases [6]Locations [7]NCT Number)

Example for Data_table:<tab sep>
--------------------------------
cancer1 cancer2 cancerN
gene1   gene    gene
gene2   gene    gene
gene3   gene    ...
gene4   ...
...
--------------------------------
"""

import os,sys,time
from urllib.request import urlopen
from urllib.error import URLError, HTTPError
from urllib import request
from urllib.request import quote
import re
from bs4 import BeautifulSoup
from openpyxl import Workbook

def clinicaltrial_Crawler(cancer1,gene_name):
    cancer = '+'.join(str(cancer1).split(' '))
    gene_name = gene_name
    url_2 = 'https://www.clinicaltrials.gov/ct2/results/download_fields?down_count=10000&down_flds=all&down_fmt=tsv&term=' + gene_name + '&cond=' + cancer + '&flds=a&flds=b&flds=f&flds=y'
    #url_1 = 'https://www.clinicaltrials.gov/ct2/results?cond=' + cancer + '&term=' + gene_name + '&cntry=&state=&city=&dist=' #https://www.clinicaltrials.gov/ct2/results?cond=Lung+Cancer&term=AKT1&cntry=&state=&city=&dist=
    print(cancer1,gene_name,': serve connecting')
    try:
        headers = {b'User-Agent': b'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.101 Safari/537.36'}
        #html = urlopen(url_1,)
        rq = request.Request(url_2,headers=headers)
        html = urlopen(rq)
        time.sleep(0.5)
    except HTTPError as e:
        if '404' in str(e):
            soup = 'not found'
            print(cancer1,gene_name,'not found')
        else:
            soup = 'none'
            print(cancer1,gene_name,'HTTP ERROR')
            print(e)
    except URLError as e:
        soup = 'none'
        print(cancer1,gene_name,'URL error')
        print(e)
    except ValueError as e:
        soup = 'none'
        print(cancer1,gene_name,'Value ERROR')
        print(e)
    except OSError as e:
        soup = 'none'
        print(cancer1,gene_name,'OSError')
        print(e)
    except Exception as e:
        soup = 'none'
        print(cancer1,gene_name,'another error:')
        print(e)
    else:
        print(cancer1,gene_name,': connected, downloading and parsing now...')
        soup = BeautifulSoup(html,"lxml")
        soup = soup.text
    finally:
        return str(soup)
        #return soup

def clinicaltrial_text1(gene_name, result_of_crawler):
    global miss_num, good_num
    count = 0
    count_g = 0
    count_b = 0
    text_line = ''
    text_orignal = ''
    if len(result_of_crawler.split('\n')) > 1:
        for i in result_of_crawler.strip().split('\n'):
            text_temp = ''
            count += 1
            # [0'Rank', 1'NCT Number', 2'Title', 3'Acronym', 4'Recruitment', 5'Study Results', 6'Conditions', 7'Interventions', 8'Outcome Measures', 9'Sponsor/Collaborators', 10'Gender', 11'Age', 12'Phases', 13'Enrollment', 14'Funded Bys', 15'Study Type', 16'Study Designs', 17'Other IDs', 18'Start Date', 19'Primary Completion Date', 20'Completion Date', 21'Last Verified', 22'First Submitted', 23'First Posted', 24'Results First Submitted', 25'Results First Posted', 26'Last Update Submitted', 27'Last Update Posted', 28'Locations', 29'URL\r']
            if count > 1:
                element = i.split('\t')
                if len(element) < 29:
                    break
                if not 'Drug' in element[7] or not 'Has' in element[5] or element[28] == '':
                    miss_num += 1
                    count_b += 1
                    #print(element)
                else:
                    try:
                        drugs = element[7].replace('|', '')
                        drugs = drugs.replace('Drug:', ',')
                        drugs = drugs.strip(',')
                        text_temp = element[6] + '\t' + gene_name + '\t' + str(drugs) + '\t' + element[2] + '\t' + element[12] + '\t' + element[28] + '\t' + element[1] + '\n'
                        text_line += text_temp
                        good_num += 1
                        count_g += 1
                        #print('effective')
                    except:
                        miss_num += 1
                        count_b += 1
            text_orignal += i + '\n'
        print('\t',gene_name,':',count_b,'line(s) deleted. Total',miss_num,'line(s) deleted.')
        print('\t',gene_name,':',count_g,'line(s) filtered. Total',good_num,'line(s) filtered.')
    else:
        print(gene_name,'clinicaltrial_text = 1')
        return 0
    #print(text_orignal)
    #print(text_line)
    return text_orignal,text_line

def single_gene(cancer,gene_name):
    a = clinicaltrial_Crawler(cancer,gene_name)
    if a == 'not found':
        print(cancer,gene_name,': has no result')
        return '',''
    elif a == 'none':
        print(cancer,gene_name,': Connection failed')
        return '',''
    else:
        t_o,t_line = clinicaltrial_text1(gene_name,a)
        return t_o,t_line

def mulit_gene(genelist):
    text_t = ''
    for i in genelist:
        text_t += '\n\nConditions\tGene\tInterventions\tTitle\tPhases\tLocations\tNCT Number\n'
        cancer = i[0]
        gene_l = i[1:]
        for j in gene_l:
            t_orignal, t_line = single_gene(cancer,j)
            text_t += t_line
    return text_t.lstrip('\n')

def excel_writing(text,output_name,sheet_name):
    text_excel = text.split('\n\n')
    Excel = Workbook()
    for i in range(len(text_excel)):
        content = text_excel[i].strip('\n').split('\n')
        content_1 = []
        for j in content:
            content_1.append(j.split('\t'))
        if i == 0:
            E_temp = Excel.active
            E_temp.title = str(sheet_name[i])
        else:
            E_temp = Excel.create_sheet(title=str(sheet_name[i]))
        for col in range(len(content_1)):
            for row in range(len(content_1[col])):
                E_temp.cell(column=row + 1, row=col + 1, value=content_1[col][row])
    Excel.save(filename=output_name)
    print('Excel file saved as: ' + output_name)

def main1(file_in):
    global miss_num, good_num
    miss_num, good_num = 0, 0
    #file_in = 'D:\\zcs-genex\\180206\\temp1.txt'
    file_in_o = open(file_in, 'r')
    names = locals()
    for i in range(10):
        names['a_%s' % str(i)] = []
    count = 0
    while 1:
        count += 1
        temp = file_in_o.readline()
        if not temp:
            break
        line = temp.strip('\n').split('\t')
        if count == 1:
            line_len = len(line)
            line_len1 = len(line)
        else:
            line_len1 = len(line)
        for i in range(line_len1):
            if line[i] != '':
                names['a_%s' % str(i)].append(line[i])
    file_in_o.close()

    genelist = []
    sheet_names = []
    for i in range(line_len):
        genelist.append(names['a_%s' % str(i)])
    print('==================================')
    for i in genelist:
        sheet_names.append(i[0])
        print(i[0], ', Gene: ', ' '.join(i[1:]))
    print('==================================')
    text = mulit_gene(genelist)
    file_out = file_in + '.out'
    file_out_o = open(file_out, 'w', encoding='utf-8')
    file_out_o.write(text)
    file_out_o.close()
    print('Tsv format file saved as: ' + file_out)

    excel_name = file_in + '.xlsx'
    excel_writing(text,excel_name,sheet_names)


if __name__ == '__main__':
    main1(sys.argv[1])





